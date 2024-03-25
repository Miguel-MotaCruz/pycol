import pandas as pd
import numpy as np
# import matplotlib.pyplot as plt
import csv
import os
from operator import itemgetter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import json

# folder_results_path = "/Users/miguel_cruz/Documents/Miguel_Cruz/LEI/CISUC/CrossValidation/src/CV_complexity/pycol/results2"


# partitioned_files = {"/Users/miguel_cruz/Library/CloudStorage/OneDrive-UniversidadedeCoimbra/CV_Kfold/arff_datasets_folds": ["housevotes.arff", "wdbc.arff", "heart.arff", "saheart.arff", "crx.arff", "haberman.arff", "spectfheart.arff", "tic-tac-toe.arff", "monk-2.arff", "pima.arff", "breast.arff", "titanic.arff", "australian.arff", "mushroom.arff", "spambase.arff", "chess.arff", "banana.arff", "sonar.arff", "bupa.arff", "phoneme.arff", "wisconsin.arff", "bands.arff", "hepatitis.arff", "german.arff", "mammographic.arff", "ionosphere.arff", "appendicitis.arff"]}
# partitioned_files = {"/Users/miguel_cruz/Library/CloudStorage/OneDrive-UniversidadedeCoimbra/CV_Kfold/arff_datasets_folds": ["titanic.arff", "housevotes.arff", "breast.arff", "sonar.arff"]}
# partitioned_files = {"/Users/miguel_cruz/Library/CloudStorage/OneDrive-UniversidadedeCoimbra/CV_Kfold/arff_datasets_folds": ["hepatitis.arff"]}
# partitioned_files = {"/Users/miguel_cruz/Documents/Miguel_Cruz/LEI/CISUC/CrossValidation/src/CV_algorithms/pyCV/new_partitions": ["caesarian-cat.arff", "hepatitis-cat.arff", "immunotherapy-cat.arff", "broadway2-cat.arff", "schizo-cat.arff", "student-g-cat.arff", "veteran-cat.arff", "traffic-cat.arff", "lymphography-v1-cat.arff", "cryotherapy-cat.arff", "servo-cat.arff", "fertility-diagnosis-cat.arff", "pharynx-1year-cat.arff", "creditscore-cat.arff", "lymphography-normal-fibrosis-cat.arff", "student-cg-cat.arff", "kidney-cat.arff", "broadwaymult0-cat.arff", "Edu-Data-HvsL-cat.arff", "pbc-cat.arff", "student-p-cat.arff", "icu-cat.arff", "cyyoung-cat.arff", "pharynx-3year-cat.arff", "heart-statlog-cat.arff", "cleveland-cat.arff", "pharynx-status-cat.arff", "broadwaymult3-cat.arff", "caesarian.arff", "cryotherapy.arff", "hepatitis.arff", "immunotherapy.arff", "creditscore.arff", "broadway3.arff", "broadway2.arff", "fertility-diagnosis.arff", "schizo.arff", "student-g.arff", "traffic.arff", "veteran.arff", "lymphography-v1.arff", "student-p.arff", "kidney.arff", "servo.arff", "lymphography-normal-fibrosis.arff", "pharynx-1year.arff", "cyyoung.arff", "pharynx-status.arff", "pharynx-3year.arff", "icu.arff", "heart-statlog.arff", "Edu-Data-HvsL.arff", "broadwaymult0.arff", "pbc.arff", "cleveland.arff", "broadwaymult6.arff", "broadwaymult4.arff", "broadwaymult3.arff", "broadwaymult5.arff", "glioma16.arff", "solvent.arff", "colon32.arff", "lupus.arff", "leukemia.arff", "appendicitis.arff", "bc-coimbra.arff", "breast-car.arff", "wine-1vs2.arff", "somerville.arff", "iris0.arff", "relax.arff", "parkinson.arff", "sonar.arff", "glass1.arff", "wpbc.arff", "ecoli_0_vs_1.arff", "newthyroid1.arff", "prnn_synth.arff", "hepato-PHvsALD.arff", "spectf.arff", "poker_9_vs_7.arff", "ecoli-0-1-3-7_vs_2-6.arff"]}
# partitioned_files = {"/Users/miguel_cruz/Documents/Miguel_Cruz/LEI/CISUC/CrossValidation/src/CV_algorithms/pyCV/new_partitions": ["ecoli_0_vs_1.arff", "poker_9_vs_7.arff", "ecoli-0-1-3-7_vs_2-6.arff"]}


def create_directory(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def join_dataset_metrics(folder, dataset_name, partition_strategy, n_splits, output_folder):
    results_path = os.path.join(folder, f"{n_splits}_folds_{partition_strategy}")
    csv_files = [f for f in os.listdir(results_path) if f.endswith('.csv') and dataset_name in f and  f'_{partition_strategy}' in f]
    df_list = [ pd.read_csv(os.path.join(results_path, csv_file), usecols=lambda column : column not in ["partition_iteration", "split"]) for csv_file in csv_files[1:]]
    if(len(csv_files)==0):
        print(results_path,dataset_name,"csv_empty")
    concatenated_df = pd.concat([pd.read_csv(os.path.join(results_path, csv_files[0]))] + df_list, axis=1)
    concatenated_df.to_csv(os.path.join(output_folder, f"{dataset_name}_{n_splits}_{partition_strategy}_all.csv"), index=False)

def join_dataset_metrics_family(folder, dataset_name, partition_strategy, n_splits, metrics, metrics_family, output_folder):
    results_path = os.path.join(folder, f"{n_splits}_folds_{partition_strategy}")
    csv_files = [f for f in os.listdir(results_path) if f.endswith('.csv') and dataset_name in f and  f"_{partition_strategy}" in f and (f.split('_')[-1].split('.')[0] if (len(f.split('_'))-len(dataset_name.split('_'))-2) == 1 else "_".join(f.split('_')[-2:]).split('.')[0]) in metrics]
    df_list = [ pd.read_csv(os.path.join(results_path, csv_file), usecols=lambda column : column not in ["partition_iteration", "split"]) for csv_file in csv_files[1:]]
    order =  ["partition_iteration", "split"] + [f"{prefix}_{metric}" for metric in metrics for prefix in ["train", "test"]]
    concatenated_df = pd.concat([pd.read_csv(os.path.join(results_path, csv_files[0]))] + df_list, axis=1)[order]
    concatenated_df.loc['mean'] = concatenated_df.mean()
    concatenated_df.loc['std'] = concatenated_df.iloc[:-1].std()
    concatenated_df.to_csv(os.path.join(output_folder, f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}.csv"), index=True)

def get_standard_deviation(folder, dataset_name, partition_strategy, n_splits, metrics_family, output_folder):
    metrics_family_file_path = os.path.join(folder, f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}.csv")
    metrics_family_file = pd.read_csv(metrics_family_file_path)
    metrics_family_file = metrics_family_file.drop(metrics_family_file.tail(2).index)
    diff_per_split = pd.DataFrame()
    for i in range(3, metrics_family_file.shape[1]-1, 2):
        diff_per_split[f"{metrics_family_file.columns[i].split('_')[1]}_dif"] = abs(metrics_family_file.iloc[:, i] - metrics_family_file.iloc[:, i+1])
    diff_per_split.loc['mean'] = diff_per_split.mean()
    diff_per_split.loc['std'] = diff_per_split.std()
    diff_per_split.to_csv(os.path.join(output_folder, f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}_std.csv"), index=True)

def concatenate_columns(folder, dataset_name, partition_strategy, n_splits, metrics_family, output_folder):
    metrics_family_file_path = os.path.join(folder, f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}.csv")
    std__file_path = os.path.join(folder, f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}_std.csv")
  
    metrics_family_df = pd.read_csv(metrics_family_file_path)
    std_df = pd.read_csv(std__file_path)
    result_df = metrics_family_df.iloc[:, :3].copy()

    for i in range(3, metrics_family_df.shape[1], 2):
        result_df = pd.concat([result_df, metrics_family_df.iloc[:, i:i+2]], axis=1)
        if (i-3) // 2 < std_df.shape[1] - 1:
            result_df = pd.concat([result_df, std_df.iloc[:, (i-3) // 2 + 1]], axis=1)
    result_df.to_csv(os.path.join(output_folder, f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}_col_concat.csv"), index=False)

# def group_std_by_partitions(dataset_name, partition_strategies, n_splits, metrics_family, input_folder, output_folder):
#     # Initialize an empty DataFrame for the result
#     result_df = pd.DataFrame()

#     # Iterate over the partition strategies
#     for partition_strategy in partition_strategies:
#         # Define the file path
#         file_path = os.path.join(input_folder, f"{n_splits}_folds_{partition_strategy}" ,f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}_col_concat.csv")
#         # Check if the file exists
#         if os.path.isfile(file_path):
#             # Read the file
#             df = pd.read_csv(file_path)
#             # drop the fisrt 3 columns
#             df = df.drop(df.columns[:3], axis=1)

#             mean_row = df.iloc[-2]
#             std_row = df.iloc[-1]
#             mean_row.name = f"{partition_strategy}_mean"
#             std_row.name = f"{partition_strategy}_std"
#             result_df = pd.concat([result_df, mean_row, std_row], axis=1)
#         else:
#             print(f"File {file_path} not found")

#     # Write the result DataFrame to a new file
#     result_df.to_csv(os.path.join(output_folder, f"{dataset_name}_{n_splits}_{metrics_family}_mean_std.csv"), index=True)

def group_std_by_partitions(dataset_name, partition_strategies, n_splits, metrics_family, input_folder, output_folder):
    # Initialize two empty DataFrames for the std and mean results
    std_df = pd.DataFrame()
    mean_df = pd.DataFrame()

    # Iterate over the partition strategies
    for partition_strategy in partition_strategies:
        # Define the file path
        file_path = os.path.join(input_folder, f"{n_splits}_folds_{partition_strategy}" ,f"{dataset_name}_{n_splits}_{partition_strategy}_{metrics_family}_col_concat.csv")
        # Check if the file exists
        if os.path.isfile(file_path):
            # Read the file
            df = pd.read_csv(file_path)
            # drop the fisrt 3 columns
            df = df.drop(df.columns[:3], axis=1)

            mean_row = df.iloc[-2]
            std_row = df.iloc[-1]
            mean_row.name = f"{partition_strategy}_mean"
            std_row.name = f"{partition_strategy}_std"
            mean_df = pd.concat([mean_df, mean_row], axis=1)
            std_df = pd.concat([std_df, std_row], axis=1)
        else:
            print(f"File {file_path} not found")

    # Concatenate the std and mean DataFrames side by side
    result_df = pd.concat([std_df, mean_df], axis=1)

    # Write the result DataFrame to a new file
    result_df.to_csv(os.path.join(output_folder, f"{dataset_name}_{n_splits}_{metrics_family}_mean_std.csv"), index=True)



def color_file(file_path):
    df = pd.read_csv(file_path)
    df.to_excel(file_path.replace(".csv", ".xlsx"), index=False, engine='openpyxl')

    # Load the workbook
    wb = load_workbook(file_path.replace(".csv", ".xlsx"))

    # Select the sheet
    ws = wb.active

    red_fill = PatternFill(start_color="FFC0C0", end_color="FFC0C0", fill_type = "solid")  # Soft red
    green_fill = PatternFill(start_color="C0FFC0", end_color="C0FFC0", fill_type = "solid")  # Soft green

    thick_border_side = Side(border_style="thick")
    thin_border_side = Side(border_style="thin")

    ws['A1'].value = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Iterate over the last three columns of both the std and mean sections
        for i in range(2, 5):
            std_cell = row[i]
            mean_cell = row[i+4]
            reference_std_cell = row[1]
            reference_mean_cell = row[5]


            if std_cell.value is  None or reference_std_cell.value is  None:
                #print the name of the dataset, column name and row name
                print(file_path.split('/')[-1], std_cell.column, std_cell.row)                

            # Compare and color the std cell
            if std_cell.value > reference_std_cell.value:
                std_cell.fill = red_fill
            elif std_cell.value < reference_std_cell.value:
                std_cell.fill = green_fill

            # Compare and color the mean cell
            if mean_cell.value > reference_mean_cell.value:
                mean_cell.fill = red_fill
            elif mean_cell.value < reference_mean_cell.value:
                mean_cell.fill = green_fill
    
    # Set the width of all columns
    for i in range(1, 10):  # Adjust the range as needed
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 15  # Set the width to 20
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter == 'A' or col_letter == 'E':  # If the cell is in the first or fifth column
                cell.border = Border(left=thin_border_side,
                                     right=thick_border_side,
                                     top=thin_border_side,
                                     bottom=thick_border_side if (cell.row - 1) % 3 == 0 else thin_border_side)
            else:
                cell.border = Border(left=thin_border_side,
                                     right=thin_border_side,
                                     top=thin_border_side,
                                     bottom=thick_border_side if (cell.row - 1) % 3 == 0 else thin_border_side)
    # Save the workbook
    wb.save(file_path.replace(".csv", ".xlsx"))



def analyze_metrics_all_datasets(dataset_names, n_splits, metrics_family, folder_results_path):
    rows_to_save = []
    # Read the first input file
    first_input_file = dataset_names[0]
    first_dataset_name = first_input_file.split(".")[0]
    first_input_file_path = os.path.join(folder_results_path, f"{first_dataset_name}_metrics",f"{first_dataset_name}_{n_splits}_{metrics_family}_mean_std.csv")
    first_input_df = pd.read_csv(first_input_file_path)
    first_input_df = first_input_df[first_input_df.iloc[:, 0].str.endswith("dif")].iloc[:, [0] + list(range(-4, 0))]

    new_df = first_input_df.copy()
    new_df.rename(columns={'Unnamed: 0': 'metrics'}, inplace=True)

    # Print the new dataframe
    count_datasets = 1
    for dataset_name in dataset_names[1:]:
        dataset_name = dataset_name.split(".")[0]
        file_path = os.path.join(folder_results_path, f"{dataset_name}_metrics",f"{dataset_name}_{n_splits}_{metrics_family}_mean_std.csv")
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            diff_rows = df[df.iloc[:, 0].str.endswith("dif")].iloc[:, [0] + list(range(-4, 0))]

            if diff_rows.isnull().values.any():
                print("\n\n!!!!!!!!!!!\n", dataset_name)
                print(diff_rows)

            new_df.iloc[:, 1:] += diff_rows.iloc[:, 1:]
            count_datasets += 1
        else:
            print(f"{file_path} does not exist")

    new_df.rename(columns={'metrics':f'metrics_{count_datasets}'}, inplace=True)
    new_df.iloc[:, 1:] /= count_datasets

    new_df.to_csv(os.path.join(folder_results_path, f"all_datasets_{n_splits}_{metrics_family}_mean.csv"), index=False)




# def save_final_metrics_to_excel(folder_results_path, metrics_families, n_splits, output_folder):
#     excel_file = os.path.join(output_folder, f"all_datasets_{n_splits}_mean.xlsx")
#     with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
#         for metrics_family in metrics_families:
#             df = pd.read_csv(os.path.join(folder_results_path, f"all_datasets_{n_splits}_{metrics_family}_mean.csv"))
#             df.to_excel(writer, sheet_name=metrics_family, index=False)
#     book = load_workbook(excel_file)
#     red_fill = PatternFill(start_color="FFC0C0", end_color="FFC0C0", fill_type = "solid")  # Soft red
#     green_fill = PatternFill(start_color="C0FFC0", end_color="C0FFC0", fill_type = "solid")  # Soft green
#     thick_border = Border(bottom=Side(style='thick'))
#     for sheet in book.sheetnames:
#         worksheet = book[sheet]
#         for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=5):
#             for cell in row:
#                 if sheet == 'multiresolution' or ('SI' in worksheet['A' + str(cell.row)].value):
#                     if cell.value < worksheet[cell.coordinate.replace(get_column_letter(cell.column), 'B')].value:
#                         cell.fill = red_fill
#                     elif cell.value > worksheet[cell.coordinate.replace(get_column_letter(cell.column), 'B')].value:
#                         cell.fill = green_fill
#                 else:
#                     if cell.value < worksheet[cell.coordinate.replace(get_column_letter(cell.column), 'B')].value:
#                         cell.fill = green_fill
#                     elif cell.value > worksheet[cell.coordinate.replace(get_column_letter(cell.column), 'B')].value:
#                         cell.fill = red_fill
#         for i in range(1, worksheet.max_column + 1):
#             worksheet.column_dimensions[get_column_letter(i)].width = 16
#             worksheet.cell(row=worksheet.max_row, column=i).border = thick_border
#     book.save(excel_file)


def save_final_metrics_to_excel(folder_results_path, metrics_families, n_splits, output_folder):
    excel_file = os.path.join(output_folder, f"all_datasets_{n_splits}_mean.xlsx")
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for metrics_family in metrics_families:
            df = pd.read_csv(os.path.join(folder_results_path, f"all_datasets_{n_splits}_{metrics_family}_mean.csv"))
            df.to_excel(writer, sheet_name=metrics_family, index=False)
    book = load_workbook(excel_file)
    red_fill = PatternFill(start_color="FFC0C0", end_color="FFC0C0", fill_type = "solid")  # Soft red
    green_fill = PatternFill(start_color="C0FFC0", end_color="C0FFC0", fill_type = "solid")  # Soft green
    thick_border = Border(bottom=Side(style='thick'))
    for sheet in book.sheetnames:
        worksheet = book[sheet]
        for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=5):
            cell_values = [cell.value for cell in row]
            max_value = max(cell_values)
            min_value = min(cell_values)
            for cell in row:
                if sheet == 'multiresolution' or ('SI' in worksheet['A' + str(cell.row)].value):
                    if cell.value == min_value:
                        cell.fill = red_fill
                    elif cell.value == max_value:
                        cell.fill = green_fill
                else:
                    if cell.value == min_value:
                        cell.fill = green_fill
                    elif cell.value == max_value:
                        cell.fill = red_fill
        for i in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 16
            worksheet.cell(row=worksheet.max_row, column=i).border = thick_border
    book.save(excel_file)

# def join_metrics_and_performance(folder_results_path, n_splits, performance_metric ,output_folder):
#     performance_file = os.path.join(folder_results_path, f"all_mean_{performance_metric}_{n_splits}.csv")
#     performance_df = pd.read_csv(performance_file)
#     mean_performance = performance_df.iloc[:, 1:].mean().tolist()
#     mean_performance = [[f"{performance_df.columns[1].split('_',1)[-1]}_{performance_metric}"] + mean_performance[:4], [f"{performance_df.columns[5].split('_',1)[-1]}_{performance_metric}"] + mean_performance[4:]]
#     excel_file = os.path.join(output_folder, f"all_datasets_{n_splits}_mean.xlsx")
#     bold_font = Font(bold=True)
#     book = load_workbook(excel_file)
#     for sheet in book.sheetnames:
#         worksheet = book[sheet]
#         for row in mean_performance:
#             worksheet.append(row)
#             worksheet.cell(row=worksheet.max_row, column=1).font = bold_font
#     book.save(excel_file)
    
# def join_metrics_and_performance(folder_results_path, n_splits, performance_metric ,output_folder):
#     performance_file = os.path.join(folder_results_path, f"all_mean_{performance_metric}_{n_splits}.csv")
#     performance_df = pd.read_csv(performance_file)
#     mean_performance = performance_df.iloc[:, 1:].mean().tolist()

#     # Calculate the number of classifiers
#     num_classifiers = len(mean_performance) // 4

#     # Create a list of lists for each classifier's mean performance
#     mean_performance_list = []
#     for i in range(num_classifiers):
#         classifier_name = performance_df.columns[i*4+1].split('_',1)[-1]
#         classifier_performance = mean_performance[i*4:(i+1)*4]
#         mean_performance_list.append([f"{classifier_name}_{performance_metric}"] + classifier_performance)

#     excel_file = os.path.join(output_folder, f"all_datasets_{n_splits}_mean.xlsx")
#     bold_font = Font(bold=True)
#     book = load_workbook(excel_file)
#     for sheet in book.sheetnames:
#         worksheet = book[sheet]
#         for row in mean_performance_list:
#             worksheet.append(row)
#             worksheet.cell(row=worksheet.max_row, column=1).font = bold_font
#     book.save(excel_file)

from openpyxl.styles import PatternFill

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def join_metrics_and_performance(folder_results_path, n_splits, performance_metric ,output_folder):
    performance_file = os.path.join(folder_results_path, f"all_mean_{performance_metric}_{n_splits}.csv")
    performance_df = pd.read_csv(performance_file)
    mean_performance = performance_df.iloc[:, 1:].mean().tolist()

    # Calculate the number of classifiers
    num_classifiers = len(mean_performance) // 4

    # Create a list of lists for each classifier's mean performance
    mean_performance_list = []
    for i in range(num_classifiers):
        classifier_name = performance_df.columns[i*4+1].split('_',1)[-1]
        classifier_performance = mean_performance[i*4:(i+1)*4]
        mean_performance_list.append([f"{classifier_name}_{performance_metric}"] + classifier_performance)

    excel_file = os.path.join(output_folder, f"all_datasets_{n_splits}_mean.xlsx")
    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color="FFC0C0", end_color="FFC0C0", fill_type = "solid")  # Soft red
    green_fill = PatternFill(start_color="C0FFC0", end_color="C0FFC0", fill_type = "solid")  # Soft green
    book = load_workbook(excel_file)
    for sheet in book.sheetnames:
        worksheet = book[sheet]
        for row in mean_performance_list:
            worksheet.append(row)
            max_value = max(row[1:])  # Exclude the classifier name
            min_value = min(row[1:])  # Exclude the classifier name
            for i in range(2, len(row) + 1):  # Start from column 2 because column 1 is the classifier name
                cell = worksheet.cell(row=worksheet.max_row, column=i)
                if cell.value == max_value:
                    cell.fill = green_fill
                elif cell.value == min_value:
                    cell.fill = red_fill
            worksheet.cell(row=worksheet.max_row, column=1).font = bold_font
    book.save(excel_file)


    

def dataset_metrics_analysis(dataset_name, partition_strategy, n_splits, feature_metrics, structural_metrics, instance_metrics, multiresolution_metrics, Metrics_folder, results_analysis_path):
    join_dataset_metrics(Metrics_folder, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, output_folder=results_analysis_path)
    join_dataset_metrics_family(Metrics_folder, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics=list(map(itemgetter(0), feature_metrics)), metrics_family="feature", output_folder=results_analysis_path)
    join_dataset_metrics_family(Metrics_folder, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics=list(map(itemgetter(0), structural_metrics)), metrics_family="structural", output_folder=results_analysis_path)
    join_dataset_metrics_family(Metrics_folder, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics=list(map(itemgetter(0), instance_metrics)), metrics_family="instance", output_folder=results_analysis_path)
    join_dataset_metrics_family(Metrics_folder, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics=list(map(itemgetter(0), multiresolution_metrics)), metrics_family="multiresolution", output_folder=results_analysis_path)

    get_standard_deviation(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="feature",output_folder=results_analysis_path)
    get_standard_deviation(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="structural",output_folder=results_analysis_path)
    get_standard_deviation(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="instance",output_folder=results_analysis_path)
    get_standard_deviation(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="multiresolution",output_folder=results_analysis_path)

    concatenate_columns(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="feature", output_folder=results_analysis_path)
    concatenate_columns(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="structural", output_folder=results_analysis_path)
    concatenate_columns(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="instance", output_folder=results_analysis_path)
    concatenate_columns(results_analysis_path, dataset_name=dataset_name, partition_strategy=partition_strategy, n_splits=n_splits, metrics_family="multiresolution", output_folder=results_analysis_path)


def dataset_mean_metrics_analysis(dataset_name, partition_strategies, n_splits, final_results_analysis_path, color_files=False):
    group_std_by_partitions(dataset_name, partition_strategies, n_splits, "feature", final_results_analysis_path, final_results_analysis_path)
    group_std_by_partitions(dataset_name, partition_strategies, n_splits, "structural", final_results_analysis_path, final_results_analysis_path)
    group_std_by_partitions(dataset_name, partition_strategies, n_splits, "instance", final_results_analysis_path, final_results_analysis_path)
    group_std_by_partitions(dataset_name, partition_strategies, n_splits, "multiresolution", final_results_analysis_path, final_results_analysis_path)
    if color_files:
        color_file(os.path.join(final_results_analysis_path, f"{dataset_name}_{n_splits}_feature_mean_std.csv"))
        color_file(os.path.join(final_results_analysis_path, f"{dataset_name}_{n_splits}_structural_mean_std.csv"))
        color_file(os.path.join(final_results_analysis_path, f"{dataset_name}_{n_splits}_instance_mean_std.csv"))
        color_file(os.path.join(final_results_analysis_path, f"{dataset_name}_{n_splits}_multiresolution_mean_std.csv"))

def final_metrics_analysis(partitioned_files, folds_folder, n_splits, folder_results_path, Performance_folder, performance_metrics):
    analyze_metrics_all_datasets(partitioned_files[folds_folder], n_splits, "feature", folder_results_path)
    analyze_metrics_all_datasets(partitioned_files[folds_folder], n_splits, "structural", folder_results_path)
    analyze_metrics_all_datasets(partitioned_files[folds_folder], n_splits, "instance", folder_results_path)
    analyze_metrics_all_datasets(partitioned_files[folds_folder], n_splits, "multiresolution", folder_results_path)
    save_final_metrics_to_excel(folder_results_path, ["feature", "structural", "instance", "multiresolution"], n_splits, folder_results_path)

    for performance_metric in performance_metrics:
        join_metrics_and_performance(Performance_folder, n_splits, performance_metric, folder_results_path)


def main(partition_strategies, feature_metrics, structural_metrics, instance_metrics, multiresolution_metrics, partitioned_files_json, Partitions_folder ,n_splits, Metrics_folder, Performance_folder, performance_metrics):
    # n_splits = 5
    # partition_strategies = ["SCV", "DBSCV", "DOBSCV", "MSSCV"]
    # feature_metrics = [('F1', 1), ('F1v', 1), ('F2', 1), ('F3', 1), ('F4', 1), ('input_noise', 1)] 
    # structural_metrics = [('N1', 0),  ('T1', 0), ('Clust', 0), ('DBC', 0), ('LSC', 0), ('NSG', 0)] #('N2', 0),
    # instance_metrics = [('R_value', 1), ('deg_overlap', 0), ('CM', 0), ('kDN', 0), ('N4', 0), ('N3', 0), ('SI', 0), ('borderline', 0)]  
    # multiresolution_metrics = [ ('C1', 0), ('C2', 0)] 
    partitioned_files = {}
    if os.path.exists(partitioned_files_json):
        if os.stat(partitioned_files_json).st_size != 0:  # Check if the file is not empty
            with open(partitioned_files_json) as json_file:
                partitioned_files = json.load(json_file)

    # for folds_folder in partitioned_files.keys():
    for dataset_name in partitioned_files[Partitions_folder]:
        dataset_name = dataset_name.split(".")[0]
        print(dataset_name)
        final_results_analysis_path = os.path.join(Metrics_folder, f"{dataset_name}_metrics")
        for partition_strategy in partition_strategies[:]:
            results_analysis_path = os.path.join(Metrics_folder, f"{dataset_name}_metrics", f"{n_splits}_folds_{partition_strategy}")
            create_directory(results_analysis_path)
            dataset_metrics_analysis(dataset_name, partition_strategy, n_splits, feature_metrics, structural_metrics, instance_metrics, multiresolution_metrics, Metrics_folder, results_analysis_path)

        dataset_mean_metrics_analysis(dataset_name, partition_strategies, n_splits, final_results_analysis_path, color_files=True)
    
    excel_file = os.path.join(Metrics_folder, f"all_datasets_{n_splits}_mean.xlsx")
    final_metrics_analysis(partitioned_files, Partitions_folder, n_splits, Metrics_folder, Performance_folder, performance_metrics)
    os.system(f"open -a 'Microsoft Excel' {excel_file}")

if __name__ == "__main__":
    main()

            
